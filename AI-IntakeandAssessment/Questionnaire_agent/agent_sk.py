import os
import asyncio
import re
import json
import openpyxl
from typing import Optional, List
from semantic_kernel import Kernel
from semantic_kernel.agents import AzureAIAgent, AzureAIAgentSettings, AzureAIAgentThread
from semantic_kernel.connectors.ai.open_ai import AzureChatCompletion
from azure.identity.aio import DefaultAzureCredential as AsyncDefaultAzureCredential
from azure.identity import AzureCliCredential
from azure.ai.projects import AIProjectClient
from azure.ai.projects.models import ConnectionType
from plugins_sk import QuestionnaireProcessorPlugin, BlobPlugin, ReviewPlugin
from tools.excel import _parse_sheet_names
from tools.telemetry import TelemetryLogger
# New imports for agent tools
from azure.ai.agents.models import (
    AzureAISearchTool,
    AzureAISearchQueryType,
    OpenApiTool,
    OpenApiAnonymousAuthDetails,
    ToolResources,
    ConnectedAgentTool,
)


class QuestionnaireAgentSK:
    """Semantic Kernel-based Questionnaire Agent with proper kernel initialization"""
    
    def __init__(self, app_id: str):
        self.app_id = app_id
        self.verbose = os.getenv("ASSESS_DEBUG", "0") in ("1", "true", "True", "yes")
        self.telemetry = TelemetryLogger()
        self.kernel: Optional[Kernel] = None
        self.agent: Optional[AzureAIAgent] = None
        self.client = None
        self._excel_path: Optional[str] = None  # Add this to track the Excel file path
        
        # Migration types mapping
        self.migration_types = {
            "1": {"name": "Modernization", "template": "ModernizationQuestionnaire.xlsx"},
            "2": {"name": "On-Prem to Azure Migration", "template": "ApplicationQuestionnaireV1.1.xlsx"},
            "3": {"name": "Cross Cloud Migration", "template": "CrossCloudQuestionnaire.xlsx"}
        }
        
        # Validate environment
        self._validate_environment()
        
    def _validate_environment(self):
        """Validate required environment variables"""
        self.endpoint = os.getenv("AZURE_AI_AGENT_ENDPOINT")
        self.model_deployment = os.getenv("AZURE_AI_AGENT_MODEL_DEPLOYMENT_NAME")
        self.search_index = os.getenv("AZURE_SEARCH_INDEX")
        # Prefer direct connection id provided by user
        self.search_connection_id = os.getenv("AZURE_SEARCH_CONNECTION_ID")
        # Optional: name fallback (not used unless you add resolver)
        self.search_connection = os.getenv("AZURE_SEARCH_CONNECTION_NAME")
        
        # Container configuration - now two separate containers
        self.templates_container = os.getenv("TEMPLATES_CONTAINER", "questionnaire-templates")
        # self.documents_container = os.getenv("DOCUMENTS_CONTAINER", f"{self.app_id}")
        documents_container_env = os.getenv("DOCUMENTS_CONTAINER", "").strip()
        self.documents_container = self.app_id
        
        # Function App config
        self.func_base_url = os.getenv("FUNC_BASE_URL")  # e.g., https://<app>.azurewebsites.net
        self.func_api_key = os.getenv("FUNC_API_KEY")
        self.target_sheets = _parse_sheet_names(os.getenv("EXCEL_SHEETS"))
        # Connected Agent B config
        self.connected_agent_b_id = os.getenv("CONNECTED_AGENT_B_ID")
        self.connected_agent_b_name = os.getenv("CONNECTED_AGENT_B_NAME") or "search_agent_b"
        self.connected_agent_b_desc = os.getenv("CONNECTED_AGENT_B_DESC") or (
            "Delegated retrieval agent that answers questions using Azure AI Search and returns compact JSON."
        )
        # Allow disabling main agent's own Search tool to force delegation
        self.disable_a_search_tool = os.getenv("DISABLE_A_SEARCH_TOOL", "0") in ("1", "true", "True", "yes")
        
        missing = []
        if not self.endpoint:
            missing.append("AZURE_AI_AGENT_ENDPOINT")
        if not self.model_deployment:
            missing.append("AZURE_AI_AGENT_MODEL_DEPLOYMENT_NAME")
        if not self.search_index:
            missing.append("AZURE_SEARCH_INDEX")
        if not self.search_connection_id and not self.connected_agent_b_id:
            # If no local search tool and no connected agent, we can't retrieve answers
            missing.append("AZURE_SEARCH_CONNECTION_ID or CONNECTED_AGENT_B_ID")
        if missing:
            raise RuntimeError(f"Missing required env vars: {', '.join(missing)}")

    
    def _vprint(self, msg: str):
        """Verbose print for debugging"""
        if self.verbose:
            print(f"[SK Agent] {msg}")
    
    async def initialize_kernel(self) -> Kernel:
        """Initialize Semantic Kernel with plugins"""
        if self.kernel:
            return self.kernel
            
        self._vprint("Initializing Semantic Kernel...")
        
        # Create kernel instance
        self.kernel = Kernel()
        
        # Add Azure OpenAI service (if using direct kernel execution)
        if os.getenv("AZURE_OPENAI_ENDPOINT") and os.getenv("AZURE_OPENAI_API_KEY"):
            service_id = "questionnaire_service"
            self.kernel.add_service(
                AzureChatCompletion(
                    service_id=service_id,
                    deployment_name=self.model_deployment,
                    endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
                    api_key=os.getenv("AZURE_OPENAI_API_KEY"),
                )
            )
            self._vprint(f"Added Azure OpenAI service: {service_id}")
        
        # Register plugins with the kernel
        self._register_plugins()
        
        self._vprint(f"Kernel initialized with {len(self.kernel.plugins)} plugins")
        return self.kernel
    
    def _register_plugins(self):
        """Register all plugins with the kernel"""
        # Create plugin instances
        processor_plugin = QuestionnaireProcessorPlugin()
        blob_plugin = BlobPlugin()
        review_plugin = ReviewPlugin()
        
        # Add plugins to kernel
        self.kernel.add_plugin(
            processor_plugin,
            plugin_name="QuestionnaireProcessor",
            description="Handles Excel operations (load, persist)"
        )
        
        self.kernel.add_plugin(
            blob_plugin,
            plugin_name="BlobOperations", 
            description="Azure Blob Storage operations"
        )
        
        self.kernel.add_plugin(
            review_plugin,
            plugin_name="ReviewManager",
            description="Manages proposal review and approval"
        )
        
        self._vprint("Registered plugins: QuestionnaireProcessor, BlobOperations, ReviewManager")
    
    async def create_agent_with_kernel(self):
        """Create Azure AI Agent with Semantic Kernel and tools (Search + Function OpenAPI)"""
        if self.agent:
            return self.agent
        
        # Initialize kernel first
        await self.initialize_kernel()
        
        # Create agent client
        self._client_creds = AsyncDefaultAzureCredential()
        self._client_cm = AzureAIAgent.create_client(
            credential=self._client_creds,
            endpoint=self.endpoint
        )
        self.client = await self._client_cm.__aenter__()
        
        # Create a separate AIProjectClient for Agent B operations if needed
        if self.connected_agent_b_id:
            from azure.ai.projects.aio import AIProjectClient
            self.project_client = AIProjectClient(
                endpoint=self.endpoint,
                credential=self._client_creds
            )
            await self.project_client.__aenter__()
        else:
            self.project_client = None
        
        # Build tools
        tools = []
        tool_objects = []
        
        # Azure AI Search tool (knowledge) for Main Agent A (optional)
        try:
            if not self.disable_a_search_tool and self.search_connection_id:
                search_tool = AzureAISearchTool(
                    index_connection_id=self.search_connection_id,
                    index_name=self.search_index,
                    query_type=AzureAISearchQueryType.SIMPLE,
                    # Default filter by appId; agent may override
                    filter=f"appId eq '{self.app_id}'",
                    top_k=50,
                )
                tool_objects.append(search_tool)
                self._vprint("Added Azure AI Search tool to main agent")
            else:
                self._vprint("Main agent Search tool disabled (using connected agent B if configured)")
        except Exception as e:
            raise RuntimeError(f"Failed to configure Azure AI Search tool: {e}")
        
        # Connected Agent B (delegated retrieval)
        if self.connected_agent_b_id:
            try:
                connected_tool = ConnectedAgentTool(
                    id=self.connected_agent_b_id,
                    name=self.connected_agent_b_name,
                    description=self.connected_agent_b_desc,
                )
                tool_objects.append(connected_tool)
                self._vprint(f"Added Connected Agent tool: {self.connected_agent_b_name} -> {self.connected_agent_b_id}")
            except Exception as e:
                raise RuntimeError(f"Failed to configure Connected Agent tool: {e}")
        
        # OpenAPI tool for Function App (inline spec hosted in code)
        if self.func_base_url and self.func_api_key:
            try:
                base = self.func_base_url.rstrip('/')

                # Build OpenAPI spec using server variables and explicit 'code' query parameter
                FUNCTION_URL = f"{base}"
                FUNCTION_KEY = self.func_api_key

                openapi_spec = {
                    "openapi": "3.0.0",
                    "info": {
                        "title": "Indexer API",
                        "version": "1.0.0",
                        "description": "Azure Function for indexing container data"
                    },
                    "servers": [
                        {
                            # embed the function key directly in the server URL so requests include the code query param
                            "url": f"{FUNCTION_URL}",
                            "variables": {
                               "functionKey": {
                                      "default": FUNCTION_KEY,
                                      "description": "Azure Function key for authentication"
                                       }
                                    }
                        }
                    ],
                    "paths": {
                        "/api/index": {
                            "post": {
                                "operationId": "indexContainer",
                                "summary": "Index a container for a given appId",
                                "parameters": [
                                        {
                                          "name": "code",
                                          "in": "query",
                                          "required": True,
                                          "schema": {
                                                "type": "string",
                                                "default": FUNCTION_KEY
                                          },
                                          "description": "Function key for authentication"
                                        }
                                ],                    
                                "requestBody": {
                                    "required": True,
                                    "content": {
                                        "application/json": {
                                            "schema": {
                                                "type": "object",
                                                "properties": {
                                                    "appId": {"type": "string"},
                                                    "container": {"type": "string"},
                                                    "blobName": {"type": ["string", "null"]}
                                                },
                                                "required": ["appId", "container"]
                                            }
                                        }
                                    }
                                },
                                "responses": {
                                    "200": {"description": "Successfully indexed"},
                                    "401": {"description": "Unauthorized - Invalid or missing API key"},
                                    "500": {"description": "Internal Server Error"}
                                }
                            }
                        }
                    }
                }

                openapi_tool = OpenApiTool(
                    name="IndexerAPI",
                    description="Azure Function that indexes container documents for the assessment.",
                    spec=openapi_spec,
                    auth=OpenApiAnonymousAuthDetails()
                )
                tool_objects.append(openapi_tool)

                # Log for debugging
                self._vprint(f"Added Function OpenAPI tool with server {FUNCTION_URL}/api/index")
            except Exception as e:
                self._vprint(f"Warning: Failed to configure Function OpenAPI tool: {e}")
                # Don't raise, just continue without the indexer tool
        else:
            self._vprint("Function tool not configured (set FUNC_BASE_URL and FUNC_API_KEY to enable)")
        
        # Convert tool objects to definitions and aggregate resources
        for t in tool_objects:
            tools.extend(t.definitions)
        # Aggregate resources (only azure_ai_search currently contributes)
        resources_dict = {}
        for t in tool_objects:
            res = getattr(t, "resources", None)
            if not res:
                continue
            if hasattr(res, "azure_ai_search") and getattr(res, "azure_ai_search"):
                resources_dict["azure_ai_search"] = getattr(res, "azure_ai_search")
        tool_resources = ToolResources(**resources_dict) if resources_dict else None
        
        # Define agent instructions
        instructions = self._get_agent_instructions()
        
        # Create agent definition with tool definitions and resources
        agent_definition = await self.client.agents.create_agent(
            model=self.model_deployment,
            name=f"SK_Questionnaire_Agent_{self.app_id}",
            instructions=instructions,
            tools=tools,
            tool_resources=tool_resources,
            headers={"x-ms-enable-preview": "true"},
        )
        
        # Create agent with kernel plugins (function tools)
        self.agent = AzureAIAgent(
            client=self.client,
            definition=agent_definition,
            kernel=self.kernel  # Pass the kernel with registered plugins
        )
        
        self._vprint(f"Created agent {self.agent.id} with kernel plugins and tools")
        return self.agent
    
    def _get_agent_instructions(self) -> str:
        """Get agent instructions with migration type handling"""
        connected_hint = (
            f"- Connected agent '{self.connected_agent_b_name}': Delegate retrieval for each question. "
            "Send only the minimal question text and filter. Expect compact JSON. Use this instead of direct search when available.\n"
            if self.connected_agent_b_id else ""
        )
        
        # Build migration type mapping for instructions
        migration_options = "\n".join([
            f"  {key}. {value['name']} -> Template: {value['template']}"
            for key, value in self.migration_types.items()
        ])
        
        return f"""
You are a Semantic Kernel-powered Questionnaire Assessment Agent.

CRITICAL RULES:
1. You MUST retrieve answers ONLY from the Azure AI Search index
2. NEVER generate, infer, or make up answers based on general knowledge
3. If no answer is found in search results, leave the field blank
4. Always use the search tool or connected agent for EVERY question

WORKFLOW:

PHASE 1: Migration Type Selection
IMPORTANT: Start by asking the user to select a migration type. Present these options:
  1. Modernization
  2. On-Prem to Azure Migraiton
  3. Cross Cloud migration

Wait for user's selection (1-3). Based on their choice, use the corresponding template:
{migration_options}

PHASE 2: Index Content
- After user selects migration type, immediately call indexContainer with appId='{self.app_id}' and container='{self.documents_container}'
- Ensure documents are indexed (uploaded >= 1)
- Success message: "The documents are indexed and ready for processing."

PHASE 3: Download Template
- Based on the migration type selected, download the appropriate template file from container '{self.templates_container}'
- Use BlobOperations.download_blob with the correct template filename
- Initialize the questionnaire with the downloaded template using QuestionnaireProcessor.initialize_questionnaire

PHASE 4: Search and Extract
- For EACH question: search the index for the answer
- Use relevant keywords from the question
- Try multiple search queries if needed
- Only use information found in search results

PHASE 5: Auto-save with Answers
- Persist the answers to Excel using QuestionnaireProcessor.persist_answers
- Upload to '{self.documents_container}' with filename format: [originalname]_filled_{self.app_id}.xlsx
- Inform user: "The answers are retrieved and ready to be evaluated. Please go and check the filled questionnaire which is uploaded in the storage account container and once done please let me know."

PHASE 6: User Review
- Wait for user to confirm they've reviewed/edited the questionnaire
- When user confirms completion, download the edited file

PHASE 7: Update Confidence
- Change all "Low" confidence values to "High"
- Change "no-results" provenance to "User-filled"
- Upload the final version
- Confirm job completion

TOOLS:
- Azure AI Search (knowledge): Search index '{self.search_index}' with filter: appId eq '{self.app_id}'
{connected_hint}- OpenAPI Function (indexContainer): Index container for the application
- Function tools (Kernel):
  - BlobOperations: download_blob, upload_file, list_blobs
  - QuestionnaireProcessor: initialize_questionnaire, persist_answers, get_excel_path, get_loaded_questions
  - ReviewManager: add_proposals, get_updates_json, clear

IMPORTANT:
- Templates are in container: {self.templates_container}
- User documents are in container: {self.documents_container}
- Start by asking for migration type selection
- Every answer must come from search results only
- Do not use your training knowledge to answer questions
- If unsure, leave blank rather than guess
"""

    async def chat_with_kernel(self, user_inputs: List[str]) -> None:
        """Chat using the kernel-based agent"""
        await self.create_agent_with_kernel()
        thread: AzureAIAgentThread = None
        
        try:
            for user_input in user_inputs:
                # Invoke agent with kernel context
                async for response in self.agent.invoke(
                    messages=user_input,
                    thread=thread,
                    kernel=self.kernel  # Pass kernel for function execution
                ):
                    print(response)
                    thread = response.thread
                    
        finally:
            await self._cleanup(thread)
    
    async def delegate_search_to_agent_b(self, question: str, row_index: int, sheet_name: str, col_index: int = 0) -> dict:
        """Delegate a single question search to Agent B with robust polling and cleanup."""
        if not self.connected_agent_b_id:
            raise RuntimeError("Connected Agent B not configured")
        
        if not self.project_client:
            raise RuntimeError("Project client not initialized")
        
        thread = None
        run = None
        try:
            # 1. Create a dedicated thread for this question
            thread = await self.project_client.agents.threads.create()
            
            search_prompt = f"""IMPORTANT: You MUST use Azure AI Search to find the answer. DO NOT generate or make up answers.

Use the {self.connected_agent_b_name} tool to search for:
Question: {question}
Filter: appId eq '{self.app_id}'

INSTRUCTIONS FOR SEARCH:
1. Search the Azure AI Search index for relevant documents.
2. Extract the answer ONLY from search results.
3. If no results found, return empty Answer.
4. DO NOT generate answers from your knowledge.
5. Return ONLY this JSON format:
{{"Answer":"<text from search results only>","Confidence":"<High if exact match, Medium if partial, Low if uncertain>","Source":"<document name from search>"}}

If search returns no results, return:
{{"Answer":"","Confidence":"Low","Source":"no-results"}}"""
            
            # 2. Add the message to the thread
            await self.project_client.agents.messages.create(
                thread_id=thread.id,
                role="user",
                content=search_prompt
            )
            
            # 3. Create the run (but don't wait for it yet)
            run = await self.project_client.agents.runs.create(
                thread_id=thread.id,
                agent_id=self.connected_agent_b_id,
                instructions="You are a search relay agent. Your ONLY job is to use the search tool and return the result in the requested JSON format. NEVER make up answers.",
                temperature=0.0,
                max_completion_tokens=500
            )

            # 4. Poll for completion with a timeout
            poll_timeout = int(os.getenv("AGENT_B_RUN_TIMEOUT", "90"))  # 90-second timeout for Agent B
            start_time = asyncio.get_event_loop().time()

            while True:
                elapsed = asyncio.get_event_loop().time() - start_time
                if elapsed > poll_timeout:
                    self._vprint(f"Run {run.id} timed out after {elapsed:.2f}s. Cancelling...")
                    raise asyncio.TimeoutError(f"Run {run.id} timed out.")

                run = await self.project_client.agents.runs.get(thread_id=thread.id, run_id=run.id)

                if run.status in ["completed", "succeeded"]:
                    break
                if run.status in ["failed", "cancelled", "canceled", "expired"]:
                    self._vprint(f"Run {run.id} terminated with status: {run.status}. Error: {run.last_error}")
                    return {
                        "RowIndex": row_index, 
                        "ColumnIndex": col_index,
                        "SheetName": sheet_name, 
                        "Question": question, 
                        "Answer": "", 
                        "Confidence": "Low", 
                        "Provenance": f"run-{run.status}"
                    }
                
                await asyncio.sleep(2) # Wait 2 seconds before polling again

            # 5. Extract result from messages
            messages = self.project_client.agents.messages.list(thread_id=thread.id, order="desc")
            async for message in messages:
                if message.role == "assistant":
                    content_text = ""
                    if message.content and isinstance(message.content, list) and hasattr(message.content[0], 'text'):
                        content_text = message.content[0].text.value
                    
                    if content_text:
                        try:
                            json_match = re.search(r'\{.*\}', content_text, re.DOTALL)
                            if json_match:
                                result = json.loads(json_match.group())
                                answer = result.get("Answer", "")
                                source = result.get("Source", "")
                                if answer and source in ["", "no-results", "unknown", None]:
                                    answer = "" # Clear potentially hallucinated answer
                                return {
                                    "RowIndex": row_index, 
                                    "ColumnIndex": col_index,
                                    "SheetName": sheet_name, 
                                    "Question": question, 
                                    "Answer": answer, 
                                    "Confidence": result.get("Confidence", "Low"), 
                                    "Provenance": source or "no-source"
                                }
                        except Exception as e:
                            self._vprint(f"Failed to parse JSON from Agent B response: {e}")
                    break # Only process the latest assistant message
            
            return {
                "RowIndex": row_index, 
                "ColumnIndex": col_index,
                "SheetName": sheet_name, 
                "Question": question, 
                "Answer": "", 
                "Confidence": "Low", 
                "Provenance": "no-response"
            }
                
        except Exception as e:
            self._vprint(f"Error in delegate_search_to_agent_b: {e}")
            return {
                "RowIndex": row_index, 
                "ColumnIndex": col_index,
                "SheetName": sheet_name, 
                "Question": question, 
                "Answer": "", 
                "Confidence": "Low", 
                "Provenance": f"error: {str(e)[:50]}"
            }
        finally:
            # 6. GUARANTEED CLEANUP
            if run and thread:
                try:
                    # Attempt to cancel the run if it's still active
                    current_run = await self.project_client.agents.runs.get(thread_id=thread.id, run_id=run.id)
                    if current_run.status not in ["completed", "succeeded", "failed", "cancelled", "canceled"]:
                        self._vprint(f"Cancelling lingering run {run.id} in finally block.")
                        await self.project_client.agents.runs.cancel(thread_id=thread.id, run_id=run.id)
                        await asyncio.sleep(1) # Give a moment for cancellation to register
                except Exception as cancel_err:
                    self._vprint(f"Could not cancel run {run.id} during cleanup: {cancel_err}")
            
            if thread:
                try:
                    self._vprint(f"Deleting thread {thread.id}")
                    await self.project_client.agents.threads.delete(thread.id)
                except Exception as delete_err:
                    self._vprint(f"Failed to delete thread {thread.id}: {delete_err}")

    async def batch_search_questions(self, questions: List[dict], batch_size: int = 5) -> List[dict]:
        """Process questions in batches through Agent B"""
        proposals = []
        
        # Process in batches with concurrency control
        semaphore = asyncio.Semaphore(3)  # Max 3 concurrent searches
        
        async def search_with_limit(q):
            async with semaphore:
                try:
                    # Pass all required parameters including ColumnIndex
                    result = await self.delegate_search_to_agent_b(
                        question=q.get("Question", ""),
                        row_index=q.get("RowIndex", 0),
                        sheet_name=q.get("SheetName", "Sheet1"),
                        col_index=q.get("ColumnIndex", 0)  # Pass the column index
                    )
                    return result
                except Exception as e:
                    self._vprint(f"Search failed for question: {e}")
                    return {
                        "RowIndex": q.get("RowIndex", 0),
                        "ColumnIndex": q.get("ColumnIndex", 0),
                        "SheetName": q.get("SheetName", "Sheet1"),
                        "Question": q.get("Question", ""),
                        "Answer": "",
                        "Confidence": "Low",
                        "Provenance": "error"
                    }
        
        # Collect ALL proposals first, then add them at once
        all_proposals = []
        
        # Process in batches without verbose output
        for i in range(0, len(questions), batch_size):
            batch_end = min(i + batch_size, len(questions))
            batch = questions[i:batch_end]
            
            # Create tasks for this batch
            tasks = [search_with_limit(q) for q in batch]
            
            # Execute batch
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            # Collect results
            for result in results:
                if isinstance(result, dict):
                    all_proposals.append(result)
                    proposals.append(result)
            
            # Small delay between batches
            await asyncio.sleep(2)
        
        # Add ALL proposals at once to the review manager
        if all_proposals:
            proposals_json = json.dumps({"proposals": all_proposals})
            await self.kernel.invoke(
                function_name="add_proposals",
                plugin_name="ReviewManager",
                proposals_json=proposals_json
            )
            self._vprint(f"Added {len(all_proposals)} proposals to ReviewManager")
        
        return proposals
    
    async def _load_question_ordinals(self):
        """Load and cache sheet -> [{Number, RowIndex, Question}] and reverse map row->Number"""
        try:
            res = await self.kernel.invoke(
                function_name="get_question_map",
                plugin_name="QuestionnaireProcessor"
            )
            text = res.value if hasattr(res, "value") else str(res)
            self._question_map = json.loads(text) if text else {}
            # build reverse map for display
            self._row_to_number = {}
            for sheet, items in (self._question_map or {}).items():
                self._row_to_number.setdefault(sheet, {})
                for it in items:
                    self._row_to_number[sheet][int(it["RowIndex"])] = int(it["Number"])
        except Exception:
            self._question_map = {}
            self._row_to_number = {}

    async def _resolve_row_by_qno(self, sheet: str, qno: int) -> Optional[dict]:
        """Resolve question number to row using plugin resolver"""
        res = await self.kernel.invoke(
            function_name="resolve_question_number",
            plugin_name="QuestionnaireProcessor",
            sheet=sheet,
            number=int(qno)
        )
        text = res.value if hasattr(res, "value") else str(res)
        try:
            j = json.loads(text) if text else {}
            return j if isinstance(j, dict) and j.get("RowIndex") else None
        except Exception:
            return None

    async def _list_unanswered(self) -> List[dict]:
        res = await self.kernel.invoke(function_name="get_unanswered", plugin_name="ReviewManager")
        text = res.value if hasattr(res, "value") else str(res)
        try:
            return json.loads(text) if text else []
        except Exception:
            return []

    async def _unanswered_count(self) -> int:
        res = await self.kernel.invoke(function_name="get_unanswered_count", plugin_name="ReviewManager")
        val = res.value if hasattr(res, "value") else str(res)
        try:
            return int(val)
        except Exception:
            return 0

    async def _set_answer_user(self, sheet: str, row_index: int, answer: str):
        await self.kernel.invoke(
            function_name="set_answer",
            plugin_name="ReviewManager",
            sheet=sheet,
            row_index=int(row_index),
            answer=answer,
            confidence="High",
            provenance="User_filled",
        )

    async def _fill_missing_interactively(self):
        """Prompt user to fill unanswered items one by one"""
        await self._load_question_ordinals()
        while True:
            unanswered = await self._list_unanswered()
            if not unanswered:
                print("All questions are answered.")
                return
            print(f"\nYou have {len(unanswered)} unanswered question(s). Type 'skip' to leave blank.")
            # show a small preview list
            preview = unanswered[:5]
            for u in preview:
                sheet = u["SheetName"]
                row = int(u["RowIndex"])
                qno = self._row_to_number.get(sheet, {}).get(row, "?")
                print(f"- {sheet} Q#{qno} (Row {row}): {u.get('Question','')[:80]}")
            # pick first
            target = unanswered[0]
            sheet = target["SheetName"]
            row = int(target["RowIndex"])
            qno = self._row_to_number.get(sheet, {}).get(row, "?")
            prompt = f"\nProvide answer for {sheet} Q#{qno} (Row {row}): "
            try:
                ans = await asyncio.get_event_loop().run_in_executor(None, lambda: input(prompt).strip())
            except Exception:
                return
            if ans.lower() == "skip":
                # leave unanswered; move on to next
                # optional: set explicit "User_filled" empty? we leave it blank
                # rotate item by moving to end (simple approach: just continue)
                continue
            await self._set_answer_user(sheet, row, ans)

    async def update_confidence_scores(self, file_path: str) -> str:
        """Update confidence scores and provenance in the reviewed questionnaire"""
        try:
            wb = openpyxl.load_workbook(file_path)
            updates_made = 0
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Find headers
                header_row = None
                confidence_col = None
                provenance_col = None
                
                # Scan for headers
                for row in range(1, min(10, ws.max_row + 1)):
                    headers = {}
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row, col).value
                        if cell_value:
                            headers[str(cell_value)] = col
                    
                    # Check for confidence and provenance columns
                    for key in headers.keys():
                        if "confidence" in key.lower():
                            confidence_col = headers[key]
                            header_row = row
                        if "provenance" in key.lower():
                            provenance_col = headers[key]
                    
                    if header_row:
                        break
                
                if not header_row:
                    continue
                
                # Update values
                for row in range(header_row + 1, ws.max_row + 1):
                    if confidence_col:
                        confidence_value = ws.cell(row, confidence_col).value
                        if confidence_value and str(confidence_value).lower() == "low":
                            ws.cell(row, confidence_col, value="High")
                            updates_made += 1
                    
                    if provenance_col:
                        provenance_value = ws.cell(row, provenance_col).value
                        if provenance_value and "no-results" in str(provenance_value).lower():
                            ws.cell(row, provenance_col, value="User-filled")
                            updates_made += 1
            
            wb.save(file_path)
            return f"Updated {updates_made} confidence/provenance values"
        except Exception as e:
            return f"Error updating confidence scores: {e}"

    async def chat_repl_with_kernel(self, initial_prompt: str = None) -> None:
        """Agent Entry Point --> Interactive REPL chat with agent using kernel and automatic search"""
        await self.create_agent_with_kernel()
        thread: AzureAIAgentThread = None
        questions_loaded = False
        search_completed = False
        
        async def invoke_with_retry(message: str, thread_obj: Optional[AzureAIAgentThread]):
            """Invoke with retry logic for rate limits and stuck runs."""
            backoff = 5
            while True:
                try:
                    # Check for and cancel any of this thread's own lingering runs before invoking
                    if thread_obj and hasattr(thread_obj, 'id'):
                        async for run in self.client.agents.runs.list(thread_id=thread_obj.id):
                            if run.status in ["in_progress", "queued", "requires_action"]:
                                self._vprint(f"Main thread has a lingering run {run.id} ({run.status}). Cancelling it.")
                                await self.client.agents.runs.cancel(thread_id=thread_obj.id, run_id=run.id)
                                await asyncio.sleep(2) # Wait for cancellation

                    async for response in self.agent.invoke(
                        messages=message,
                        thread=thread_obj,
                        kernel=self.kernel
                    ):
                        print(response)
                        return response.thread
                except Exception as e:
                    msg = str(e)
                    if "Can't add messages to thread" in msg and "while a run" in msg:
                        self._vprint(f"Caught a blocked thread error: {msg}. Retrying after a delay.")
                        await asyncio.sleep(backoff)
                        continue # The loop will re-check and cancel the lingering run

                    m = re.search(r"Try again in\s*([0-9]+)\s*seconds", msg, re.IGNORECASE)
                    wait = int(m.group(1)) if m else backoff
                    self._vprint(f"Rate limit or other error: {e}. Waiting {wait}s...")
                    await asyncio.sleep(wait)
                    backoff = min(backoff * 2, 30)
        
        async def check_and_process_questions():
            """Check if questions are loaded and process them through Agent B"""
            nonlocal questions_loaded, search_completed
            
            if questions_loaded or search_completed:
                return False
            
            try:
                # Check if questions are loaded
                questions_json = await self.kernel.invoke(
                    function_name="get_loaded_questions",
                    plugin_name="QuestionnaireProcessor"
                )
                questions = json.loads(str(questions_json.value if hasattr(questions_json, 'value') else questions_json)) if questions_json else []
                
                if questions and len(questions) > 0:
                    questions_loaded = True
                    
                    # Perform batch search silently
                    if self.connected_agent_b_id:
                        await self.batch_search_questions(questions, batch_size=5)
                        search_completed = True
                        
                        # Auto-save and upload
                        await self.save_with_kernel(self.documents_container)
                        
                        print("\n✅ The answers are retrieved and ready to be evaluated.")
                        print("Please go and check the filled questionnaire which is uploaded in the storage account container.")
                        print("Once done, please let me know by typing 'done' or 'reviewed'.")
                        return True
                    else:
                        print("⚠️ Agent B not configured. Please set CONNECTED_AGENT_B_ID environment variable.")
                        return False
                        
            except Exception as e:
                # Questions not yet loaded, continue normally
                return False
            
            return False
        
        try:
            # Start with initial prompt to begin the workflow
            initial_message = initial_prompt or "Hello! I'm ready to help you with the questionnaire assessment. Let's begin."
            thread = await invoke_with_retry(initial_message, thread)
            
            # Interactive loop
            while True:
                # Check if we should auto-process questions
                if await check_and_process_questions():
                    # Wait for user review confirmation
                    pass
                
                try:
                    user_input = await asyncio.get_event_loop().run_in_executor(
                        None, 
                        lambda: input("You: ").strip()
                    )
                except Exception:
                    break
                    
                if not user_input or user_input.lower() in ("exit", "quit"):
                    break
                
                # Check if user has reviewed the questionnaire
                if user_input.lower() in ("done", "reviewed", "completed") and search_completed:
                    print("\nDownloading the reviewed questionnaire to update confidence scores...")
                    
                    # Download the filled questionnaire
                    if self._excel_path:
                        base_name = os.path.splitext(os.path.basename(self._excel_path))[0]
                        blob_name = f"{base_name}_filled_{self.app_id}.xlsx"
                    else:
                        # Fallback: try to get it from the plugin
                        path_result = await self.kernel.invoke(
                            function_name="get_excel_path",
                            plugin_name="QuestionnaireProcessor"
                        )
                        excel_path = path_result.value if hasattr(path_result, 'value') else str(path_result)
                        if excel_path:
                            base_name = os.path.splitext(os.path.basename(excel_path))[0]
                            blob_name = f"{base_name}_filled_{self.app_id}.xlsx"
                        else:
                            # Last resort: use a generic name based on migration type
                            blob_name = f"ApplicationQuestionnaireV1.1.xlsx.clean_filled_{self.app_id}.xlsx"
                    
                    self._vprint(f"Attempting to download: {blob_name} from container: {self.documents_container}")
                    
                    local_path = await self.kernel.invoke(
                        function_name="download_blob",
                        plugin_name="BlobOperations",
                        container=self.documents_container,
                        blob_name=blob_name
                    )
                    local_path = local_path.value if hasattr(local_path, 'value') else str(local_path)
                    
                    # Update confidence scores
                    update_result = await self.update_confidence_scores(local_path)
                    print(f"Update result: {update_result}")
                    
                    # Upload final version
                    base = os.path.splitext(os.path.basename(local_path))[0]
                    final_name = f"{base.replace('_filled', '')}_final_{self.app_id}.xlsx"
                    
                    upload_result = await self.kernel.invoke(
                        function_name="upload_file",
                        plugin_name="BlobOperations",
                        container=self.documents_container,
                        local_path=local_path,
                        dest_blob_name=final_name
                    )
                    
                    print(f"\n✅ Job completed! Final questionnaire uploaded as: {final_name}")
                    break

                # Default: pass to agent
                thread = await invoke_with_retry(user_input, thread)
                
        finally:
            await self._cleanup(thread)

    async def save_with_kernel(self, container_default: Optional[str] = None):
        try:
            # 1) Get updates (local, no LLM) - extract the value properly
            updates_result = await self.kernel.invoke(
                function_name="get_updates_json",
                plugin_name="ReviewManager"
            )
            # Extract the actual value from the FunctionResult
            updates_json = updates_result.value if hasattr(updates_result, 'value') else str(updates_result)
            
            # Debug: Check what we got
            self._vprint(f"Got updates type: {type(updates_json)}, length: {len(updates_json) if isinstance(updates_json, str) else 'N/A'}")
            
            # Parse to verify we have actual data
            try:
                updates_list = json.loads(updates_json) if isinstance(updates_json, str) else updates_json
                self._vprint(f"Updates to persist: {len(updates_list)} answers")
                if len(updates_list) == 0:
                    self._vprint("WARNING: No updates to save! Check if proposals were added correctly.")
                    return
            except Exception as e:
                self._vprint(f"ERROR parsing updates: {e}")
                self._vprint(f"Raw updates_json: {updates_json[:500] if isinstance(updates_json, str) else updates_json}")
                return

            # 2) Persist to Excel (local, no LLM) - pass the actual JSON string
            persist_result = await self.kernel.invoke(
                function_name="persist_answers",
                plugin_name="QuestionnaireProcessor",
                updates_json=updates_json  # Pass the actual JSON string, not str() of the result object
            )
            persist_status = persist_result.value if hasattr(persist_result, 'value') else str(persist_result)
            self._vprint(f"Persist status: {persist_status}")

            # 3) Get path and upload
            path_result = await self.kernel.invoke(
                function_name="get_excel_path",
                plugin_name="QuestionnaireProcessor"
            )
            local_path = path_result.value if hasattr(path_result, 'value') else str(path_result)
            
            if not local_path or local_path == "":
                self._vprint("ERROR: No Excel path available. Excel may not be initialized.")
                return
                
            self._vprint(f"Excel path: {local_path}")
            
            # Cache the excel path at class level for later use
            self._excel_path = local_path
            
            # Check if file exists and has been modified
            if os.path.exists(local_path):
                file_size = os.path.getsize(local_path)
                self._vprint(f"File exists at {local_path}, size: {file_size} bytes")
            else:
                self._vprint(f"WARNING: File not found at {local_path}")
                return
            
            base = os.path.splitext(os.path.basename(local_path))[0]
            dest_name = f"{base}_filled_{self.app_id}.xlsx"
            
            # Use the provided container or the configured documents container
            container = container_default or self.documents_container
            
            # Validate container name
            if not container or len(container) < 3 or len(container) > 63:
                self._vprint(f"Invalid container name: {container}. Using app_id as container.")
                container = self.app_id
                
            self._vprint(f"Uploading to container: {container} as {dest_name}")
            
            upload_result = await self.kernel.invoke(
                function_name="upload_file",
                plugin_name="BlobOperations",
                container=container,
                local_path=local_path,
                dest_blob_name=dest_name
            )
            upload_url = upload_result.value if hasattr(upload_result, 'value') else str(upload_result)
            self._vprint(f"Saved and uploaded: {upload_url}")
            
        except Exception as e:
            self._vprint(f"ERROR in save_with_kernel: {e}")
            import traceback
            traceback.print_exc()

    async def _cleanup(self, thread: Optional[AzureAIAgentThread]):
        """Clean up resources"""
        try:
            if thread:
                await thread.delete()
        except Exception:
            pass
            
        try:
            if self.agent and self.client:
                agent_id = getattr(self.agent, 'id', None)
                if agent_id:
                    await self.client.agents.delete_agent(agent_id)
        except Exception:
            pass
        
        try:
            if hasattr(self, "project_client") and self.project_client:
                await self.project_client.__aexit__(None, None, None)
        except Exception:
            pass
            
        try:
            if hasattr(self, "_client_cm") and self._client_cm:
                await self._client_cm.__aexit__(None, None, None)
        except Exception:
            pass
            
        try:
            if hasattr(self, "_client_creds") and self._client_creds:
                await self._client_creds.close()
        except Exception:
            pass
    
    async def close(self):
        """Close all resources"""
        await self._cleanup(None)
