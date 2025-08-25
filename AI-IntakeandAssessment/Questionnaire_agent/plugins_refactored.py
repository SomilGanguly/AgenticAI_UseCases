import os, json
import openpyxl
from typing import Annotated, List, Dict, Optional, Tuple
from semantic_kernel.functions import kernel_function

from tools import blob as blob_utils
from tools.xlsx_zip import materialize_clean_xlsx


QUESTION_VARIATIONS = [
    "Question", "Questions", "question", "questions", "Query", "query", "Q", "q"
]
GUIDANCE_VARIATIONS = [
    "Guidance", "guidance", "Examples", "examples", "Instructions", "instructions", "Help", "help", "Guide", "guide"
]
ANSWER_VARIATIONS = [
    "Response", "Responses", "response", "responses", "Answer", "Answers", "answer", "answers", "Reply", "reply", "A", "a"
]
TRACKING_COLUMNS = ["Confidence", "Provenance"]


def _parse_sheets_csv(sheets_csv: str) -> Optional[List[str]]:
    if not sheets_csv:
        return None
    return [s.strip() for s in sheets_csv.split(',') if s and s.strip()]


def _header_index(ws, header_row: int = 1) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    if header_row < 1 or header_row > (ws.max_row or 1):
        return headers
    for idx, cell in enumerate(ws[header_row], start=1):
        if cell.value is not None:
            headers[str(cell.value)] = idx
    return headers


def _find_column_case_insensitive(headers: Dict[str, int], variations: List[str]) -> Optional[str]:
    for var in variations:
        if var in headers:
            return var
    headers_lower = {k.lower(): k for k in headers.keys() if k is not None}
    for var in variations:
        if var.lower() in headers_lower:
            return headers_lower[var.lower()]
    for header in headers.keys():
        if header is None:
            continue
        hl = header.lower()
        for var in variations:
            if var.lower() in hl:
                return header
    return None


def _find_or_create_column(ws, headers: Dict[str, int], variations: List[str], default_name: str, header_row: int = 1) -> Tuple[Dict[str, int | str], str]:
    # Prefer an existing column by variations first
    found = _find_column_case_insensitive(headers, variations)
    if found:
        return headers, found
    # Create at true end of sheet (not len(headers)+1 which can land inside merged regions)
    new_col_idx = ws.max_column + 1
    ws.cell(header_row, new_col_idx, value=default_name)
    headers[default_name] = new_col_idx
    return headers, default_name


def _scan_for_question_header(ws, max_scan_rows: int = 10) -> Tuple[int, Optional[int], Optional[str]]:
    """Scan for the row containing actual column headers (Guidance, Response, etc.)"""
    max_row = max(1, ws.max_row or 1)
    scan_to = min(max_scan_rows, max_row)
    
    # Strategy 1: Look for a row with Guidance AND Response columns (most reliable)
    for r in range(1, scan_to + 1):
        headers = _header_index(ws, header_row=r)
        
        # Check if this row has the key headers we expect
        has_guidance = _find_column_case_insensitive(headers, GUIDANCE_VARIATIONS) is not None
        has_response = _find_column_case_insensitive(headers, ANSWER_VARIATIONS) is not None
        
        # If we find both Guidance and Response, this is definitely our header row
        if has_guidance and has_response:
            # The question column is likely the first non-empty column before Guidance
            # or it might not have a header at all
            q_col_name = _find_column_case_insensitive(headers, QUESTION_VARIATIONS)
            if q_col_name:
                return r, headers[q_col_name], q_col_name
            else:
                # No explicit Question header, assume column 2 (B) has questions
                # (column 1 might be merged title)
                return r, 2, None
    
    # Strategy 2: Look for a row with at least Response column
    for r in range(1, scan_to + 1):
        headers = _header_index(ws, header_row=r)
        has_response = _find_column_case_insensitive(headers, ANSWER_VARIATIONS) is not None
        
        if has_response and len(headers) > 2:  # Multiple headers suggest this is the header row
            q_col_name = _find_column_case_insensitive(headers, QUESTION_VARIATIONS)
            if q_col_name:
                return r, headers[q_col_name], q_col_name
            else:
                # No Question header, use column 2
                return r, 2, None
    
    # Strategy 3: Fallback - look for Question column
    for r in range(1, scan_to + 1):
        headers = _header_index(ws, header_row=r)
        q = _find_column_case_insensitive(headers, QUESTION_VARIATIONS)
        if q:
            return r, headers[q], q
    
    # Last resort: assume header is in row 3 (common in templates) and questions in column 2
    return 3, 2, None

class QuestionnaireProcessorPlugin:
    """
    Handles Excel operations only (load questions, persist answers). Search is delegated to the agent tools.
    """
    
    def __init__(self):
        self._loaded_questions: List[Dict] = []
        self._excel_path: str = ""
        self._column_maps: Dict[str, Dict[str, int | str]] = {}
        self._selected_sheets: List[str] = []
        
    @kernel_function(description="Initialize questionnaire processing by loading Excel and preparing for Q&A")
    def initialize_questionnaire(
        self,
        path: Annotated[str, "Local filesystem path to the Excel workbook"],
        sheets_csv: Annotated[str, "Comma-separated sheet names to process; allow multiple; if empty, uses active sheet"] = "",
    ) -> Annotated[str, "Status message with question count"]:
        try:
            clean_path = materialize_clean_xlsx(path)
            wb = openpyxl.load_workbook(clean_path)
            sheets = _parse_sheets_csv(sheets_csv) or [wb.active.title]
            self._selected_sheets = sheets

            self._loaded_questions = []
            self._column_maps = {}

            changed = False
            for sheet_name in sheets:
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                header_row, q_col_idx, _ = _scan_for_question_header(ws)
                
                # Debug output
                print(f"Sheet '{sheet_name}': Found header row at {header_row}, question column at {q_col_idx}")
                
                headers = _header_index(ws, header_row=header_row)
                print(f"Headers found: {list(headers.keys())}")

                # Ensure Answer column exists
                original_headers = headers.copy()
                headers, answer_col_name = _find_or_create_column(ws, headers, ANSWER_VARIATIONS, "Response", header_row)
                if headers != original_headers:
                    changed = True
                answer_col_idx = headers[answer_col_name]

                # Optional guidance column detection
                guidance_col_name = _find_column_case_insensitive(headers, GUIDANCE_VARIATIONS)
                guidance_col_idx = headers.get(guidance_col_name) if guidance_col_name else None

                if not q_col_idx:
                    # fallback: assume column 2 (B) is questions (column 1 might be title/merged)
                    q_col_idx = 2

                self._column_maps[sheet_name] = {
                    "header_row": header_row,
                    "question_col": q_col_idx,
                    "answer_col": answer_col_idx,
                    "guidance_col": guidance_col_idx or 0,
                }

                # Load questions - start from the row after the header
                start_row = header_row + 1
                max_row = ws.max_row or start_row
                
                print(f"Loading questions from rows {start_row} to {max_row}, column {q_col_idx}")
                
                for r in range(start_row, max_row + 1):
                    qv = ws.cell(r, q_col_idx).value
                    if qv is None or str(qv).strip() == "":
                        continue
                    guidance_val = ws.cell(r, guidance_col_idx).value if guidance_col_idx else None
                    self._loaded_questions.append({
                        "RowIndex": r,
                        "SheetName": sheet_name,
                        "Question": str(qv),
                        "Guidance": str(guidance_val) if guidance_val is not None else "",
                    })

            if changed:
                wb.save(clean_path)
            self._excel_path = clean_path
            return f"Loaded {len(self._loaded_questions)} questions from {len(self._column_maps)} sheet(s)."
        except Exception as e:
            import traceback
            traceback.print_exc()
            return f"Error: {e}"

    @kernel_function(description="Save approved answers back to Excel")
    def persist_answers(
        self,
        updates_json: Annotated[str, "JSON array of updates with RowIndex, SheetName, Answer, Confidence, Provenance"],
    ) -> Annotated[str, "Status message"]:
        if not self._excel_path:
            return "Error: Excel not initialized"
        try:
            updates = json.loads(updates_json) if isinstance(updates_json, str) else updates_json
            if not isinstance(updates, list):
                return "Error: updates_json must be a JSON array"

            wb = openpyxl.load_workbook(self._excel_path)

            # Helper to write to data cells (handles merged cells properly)
            def write_data_cell(ws, row, col, value):
                """Write to a specific cell, handling merged ranges"""
                if col <= 0:
                    return  # Invalid column
                    
                # First, check if this cell is part of a merged range
                merged_ranges_to_unmerge = []
                for merged_range in ws.merged_cells.ranges:
                    if row >= merged_range.min_row and row <= merged_range.max_row and \
                       col >= merged_range.min_col and col <= merged_range.max_col:
                        merged_ranges_to_unmerge.append(str(merged_range))
                
                # Unmerge any ranges that contain this cell
                for range_str in merged_ranges_to_unmerge:
                    ws.unmerge_cells(range_str)
                
                # Now write the value
                ws.cell(row, col, value=value)

            # Process each sheet to ensure columns exist and get their indices
            sheet_column_info = {}
            for sheet_name in self._column_maps.keys():
                if sheet_name not in wb.sheetnames:
                    continue
                    
                ws = wb[sheet_name]
                # Use the stored header_row from initialization
                header_row = int(self._column_maps[sheet_name]["header_row"])
                
                print(f"\nProcessing sheet '{sheet_name}' with header row: {header_row}")
                
                # Get current headers from the correct row
                headers = _header_index(ws, header_row)
                print(f"Headers at row {header_row}: {list(headers.keys())}")
                
                # Find Response/Answer column
                answer_col_name = _find_column_case_insensitive(headers, ANSWER_VARIATIONS)
                if answer_col_name:
                    answer_col_idx = headers[answer_col_name]
                    print(f"Found existing '{answer_col_name}' column at index: {answer_col_idx}")
                else:
                    # Create Response column at the end, in the same header row
                    new_col_idx = ws.max_column + 1
                    print(f"Creating 'Response' column at index: {new_col_idx} in row {header_row}")
                    ws.cell(header_row, new_col_idx, value="Response")
                    # Refresh headers after adding
                    headers = _header_index(ws, header_row)
                    answer_col_name = "Response"
                    answer_col_idx = new_col_idx
                
                # Find or create Confidence column (optional - only if needed)
                confidence_col_idx = headers.get("Confidence", 0)
                if not confidence_col_idx and any(upd.get("Confidence") for upd in updates if upd.get("SheetName") == sheet_name):
                    new_col_idx = ws.max_column + 1
                    print(f"Creating 'Confidence' column at index: {new_col_idx} in row {header_row}")
                    ws.cell(header_row, new_col_idx, value="Confidence")
                    headers = _header_index(ws, header_row)
                    confidence_col_idx = headers.get("Confidence", 0)
                
                # Find or create Provenance column (optional - only if needed)
                provenance_col_idx = headers.get("Provenance", 0)
                if not provenance_col_idx and any(upd.get("Provenance") for upd in updates if upd.get("SheetName") == sheet_name):
                    new_col_idx = ws.max_column + 1
                    print(f"Creating 'Provenance' column at index: {new_col_idx} in row {header_row}")
                    ws.cell(header_row, new_col_idx, value="Provenance")
                    headers = _header_index(ws, header_row)
                    provenance_col_idx = headers.get("Provenance", 0)
                
                # Store the column indices for this sheet
                sheet_column_info[sheet_name] = {
                    "answer_col": answer_col_idx,
                    "confidence_col": confidence_col_idx,
                    "provenance_col": provenance_col_idx,
                    "header_row": header_row
                }
                
                print(f"Final column mapping - Answer: {answer_col_idx}, Confidence: {confidence_col_idx}, Provenance: {provenance_col_idx}")

            # Now apply updates to the correct columns
            updates_applied = 0
            for upd in updates:
                sheet = upd.get("SheetName")
                row = int(upd.get("RowIndex")) if upd.get("RowIndex") is not None else None
                
                if not sheet or row is None:
                    continue
                if sheet not in wb.sheetnames:
                    continue
                if sheet not in sheet_column_info:
                    continue
                    
                ws = wb[sheet]
                col_info = sheet_column_info[sheet]
                
                # Get the values to write
                answer = upd.get("Answer", "")
                confidence = upd.get("Confidence", "")
                provenance = upd.get("Provenance", "")
                
                # Write to the Response/Answer column
                answer_col = col_info["answer_col"]
                if answer_col and answer_col > 0:
                    if updates_applied < 5:  # Only print first few for debugging
                        print(f"Writing to row {row}, Response column {answer_col}: '{answer[:30]}...'")
                    write_data_cell(ws, row, answer_col, answer)
                else:
                    print(f"WARNING: No answer column found for sheet {sheet}")
                    
                # Optionally write confidence and provenance (only if columns exist and values are non-empty)
                if col_info["confidence_col"] > 0 and confidence:
                    write_data_cell(ws, row, col_info["confidence_col"], confidence)
                    
                if col_info["provenance_col"] > 0 and provenance:
                    write_data_cell(ws, row, col_info["provenance_col"], provenance)
                
                updates_applied += 1

            wb.save(self._excel_path)
            print(f"\nSuccessfully saved {updates_applied} answers to Excel at {self._excel_path}")
            return f"Saved {updates_applied} answers to Excel."
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return f"Error: {e}"

    @kernel_function(description="Get current Excel file path for upload")
    def get_excel_path(self) -> Annotated[str, "Path to the processed Excel file"]:
        return self._excel_path or ""

    @kernel_function(description="Return loaded questions as JSON array")
    def get_loaded_questions(self) -> Annotated[str, "JSON array of loaded questions with RowIndex, SheetName, Question, Guidance"]:
        """Get all loaded questions as JSON array for processing"""
        return json.dumps(self._loaded_questions or [])


# Keep lightweight plugins separate
class BlobPlugin:
    """Lightweight plugin for blob operations"""
    
    @kernel_function(description="Download a blob from Azure Storage")
    def download_blob(
        self,
        container: Annotated[str, "Azure Storage container name"],
        blob_name: Annotated[str, "Name of the blob to download"],
    ) -> Annotated[str, "Local file path to the downloaded blob"]:
        return blob_utils.download_blob(container, blob_name)
    
    @kernel_function(description="Upload a file to Azure Storage")
    def upload_file(
        self,
        container: Annotated[str, "Azure Storage container name"],
        local_path: Annotated[str, "Local filesystem path to upload"],
        dest_blob_name: Annotated[str, "Destination blob name"] = "",
    ) -> Annotated[str, "URL of the uploaded blob"]:
        return blob_utils.upload_file(container, local_path, dest_blob_name or None)
    
    @kernel_function(description="List blobs in a container")
    def list_blobs(
        self,
        container: Annotated[str, "Azure Storage container name"],
        prefix: Annotated[str, "Optional prefix filter"] = "",
    ) -> Annotated[str, "JSON array of blob names"]:
        blobs = blob_utils.list_blobs(container, prefix or None)
        return json.dumps(blobs)


class ReviewPlugin:
    """Lightweight plugin for review operations"""
    
    def __init__(self):
        self._proposals: List[Dict] = []
        self._title: str = ""
    
    @kernel_function(description="Get current proposal count for debugging")
    def get_proposal_count(self) -> Annotated[str, "Number of proposals currently stored"]:
        """Debug function to check if proposals are preserved"""
        return str(len(self._proposals))
    
    @kernel_function(description="Add proposals from processing (upsert by SheetName+RowIndex)")
    def add_proposals(
        self,
        proposals_json: Annotated[str, "JSON with proposals array or object with 'proposals' key"],
    ) -> Annotated[str, "Status with total count"]:
        try:
            data = json.loads(proposals_json) if isinstance(proposals_json, str) else proposals_json
            items = data.get("proposals") if isinstance(data, dict) else data
            if not isinstance(items, list):
                return "Error: proposals_json must be an array or { 'proposals': [] }"
            # upsert by (SheetName, RowIndex)
            by_key: Dict[Tuple[str, int], Dict] = {}
            for p in self._proposals:
                key = (p.get("SheetName", ""), int(p.get("RowIndex", 0)))
                by_key[key] = p
            for it in items:
                sheet = it.get("SheetName", "")
                row = int(it.get("RowIndex", 0))
                key = (sheet, row)
                by_key[key] = {
                    "SheetName": sheet,
                    "RowIndex": row,
                    "Question": it.get("Question", ""),
                    "Answer": it.get("Answer", ""),
                    "Confidence": it.get("Confidence", "Unknown"),
                    "Provenance": it.get("Provenance", ""),
                }
            self._proposals = list(by_key.values())
            return f"Total proposals: {len(self._proposals)}"
        except Exception as e:
            return f"Error: {e}"
    
    @kernel_function(description="Render proposals as Markdown table")
    def render_markdown(self) -> Annotated[str, "Markdown table"]:
        if not self._proposals:
            return "No proposals to review."
        # sort for stable output
        rows = sorted(self._proposals, key=lambda x: (x.get("SheetName", ""), int(x.get("RowIndex", 0))))
        header = "| # | Sheet | Row | Question | Answer | Confidence | Provenance |\n|---|---|---:|---|---|---|---|"
        lines = [header]
        for i, p in enumerate(rows, start=1):
            q = (p.get("Question", "") or "").replace("\n", " ")
            a = (p.get("Answer", "") or "").replace("\n", " ")
            conf = p.get("Confidence", "Unknown") or "Unknown"
            prov = (p.get("Provenance", "") or "").replace("\n", " ")
            lines.append(f"| {i} | {p.get('SheetName','')} | {p.get('RowIndex','')} | {q} | {a} | {conf} | {prov} |")
        return "\n".join(lines)
    
    @kernel_function(description="Get approved proposals as JSON (RowIndex, SheetName, Answer, Confidence, Provenance)")
    def get_updates_json(self) -> Annotated[str, "JSON array of updates"]:
        updates = []
        for p in self._proposals:
            updates.append({
                "SheetName": p.get("SheetName", ""),
                "RowIndex": p.get("RowIndex", 0),
                "Answer": p.get("Answer", ""),
                "Confidence": p.get("Confidence", "Unknown"),
                "Provenance": p.get("Provenance", ""),
            })
        return json.dumps(updates)
    
    @kernel_function(description="Clear all proposals")
    def clear(self) -> Annotated[str, "Status"]:
        self._proposals = []
        self._title = ""
        return "Cleared"