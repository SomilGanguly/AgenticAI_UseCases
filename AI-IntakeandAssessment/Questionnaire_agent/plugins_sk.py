"""
Overview:
- QuestionnaireProcessorPlugin: Excel loading, detection (row vs column), and persistence of answers.
- BlobPlugin: Minimal Azure Storage helpers (download/upload/list) used by the agent via SK.
- ReviewPlugin: Holds proposals/updates between retrieval and persistence; provides review utilities.

"""

import os, json
import openpyxl
from typing import Annotated, List, Dict, Optional, Tuple
from semantic_kernel.functions import kernel_function

from tools import blob as blob_utils
from tools.xlsx_zip import materialize_clean_xlsx


QUESTION_VARIATIONS = [
    "Question", "Questions", "question", "questions"
]
GUIDANCE_VARIATIONS = [
    "Guidance", "guidance", "Examples", "examples", "Instructions", "instructions", "Help", "help", "Guide", "guide"
]
ANSWER_VARIATIONS = [
    "Response", "Responses", "response", "responses", "Answer", "Answers", "answer", "answers", "Reply", "reply", "A", "a"
]
TRACKING_COLUMNS = ["Confidence", "Provenance"]


def _parse_sheets_csv(sheets_csv: str) -> Optional[List[str]]:
    """Split a comma-separated list of sheet names into a list (or None)."""
    if not sheets_csv:
        return None
    return [s.strip() for s in sheets_csv.split(',') if s and s.strip()]


def _header_index(ws, header_row: int = 1) -> Dict[str, int]:
    """Build a mapping of header text -> column index for a given row."""
    headers: Dict[str, int] = {}
    if header_row < 1 or header_row > (ws.max_row or 1):
        return headers
    for idx, cell in enumerate(ws[header_row], start=1):
        if cell.value is not None:
            headers[str(cell.value)] = idx
    return headers


def _find_column_case_insensitive(headers: Dict[str, int], variations: List[str]) -> Optional[str]:
    """Find a header that matches any variation (case-insensitive, substring fallback)."""
    for var in variations:
        if var in headers:
            return var
    headers_lower = {k.lower(): k for k in headers.keys() if k is not None}
    for var in variations:
        if var.lower() in headers_lower:
            return headers_lower[var.lower()]
    for header in headers.keys():
        if header is None:
            continue
        hl = header.lower()
        for var in variations:
            if var.lower() in hl:
                return header
    return None


def _find_or_create_column(ws, headers: Dict[str, int], variations: List[str], default_name: str, header_row: int = 1) -> Tuple[Dict[str, int | str], str]:
    """Ensure a specific header column exists; create if missing (appends to the right)."""
    # Prefer an existing column by variations first
    found = _find_column_case_insensitive(headers, variations)
    if found:
        return headers, found
    
    # Find the rightmost non-empty column to avoid placing headers in the middle of data
    max_col = 1
    for row in range(1, min(10, ws.max_row + 1)):
        for col in range(ws.max_column, 0, -1):
            if ws.cell(row, col).value is not None:
                max_col = max(max_col, col)
                break
    
    # Create new column after the last non-empty column
    new_col_idx = max_col + 1
    
    # Ensure we write the header value
    cell = ws.cell(header_row, new_col_idx)
    cell.value = default_name
    
    # Add to headers dictionary
    headers[default_name] = new_col_idx
    
    return headers, default_name


def _scan_for_question_header(ws, max_scan_rows: int = 10) -> Tuple[int, Optional[int], Optional[str]]:
    """Locate the header row for row-based sheets and return (row, question_col_index, question_header_text)."""
    max_row = max(1, ws.max_row or 1)
    scan_to = min(max_scan_rows, max_row)
    
    # Strategy 1: Look for a row with Guidance AND Response columns (most reliable)
    for r in range(1, scan_to + 1):
        headers = _header_index(ws, header_row=r)
        
        # Check if this row has the key headers we expect
        has_guidance = _find_column_case_insensitive(headers, GUIDANCE_VARIATIONS) is not None
        has_response = _find_column_case_insensitive(headers, ANSWER_VARIATIONS) is not None
        
        # If we find both Guidance and Response, this is definitely our header row
        if has_guidance and has_response:
            q_col_name = _find_column_case_insensitive(headers, QUESTION_VARIATIONS)
            if q_col_name:
                return r, headers[q_col_name], q_col_name
            else:
                return r, 2, None
    
    # Strategy 2: Look for a row with at least Response column
    for r in range(1, scan_to + 1):
        headers = _header_index(ws, header_row=r)
        has_response = _find_column_case_insensitive(headers, ANSWER_VARIATIONS) is not None
        
        if has_response and len(headers) > 2:  # Multiple headers suggest this is the header row
            q_col_name = _find_column_case_insensitive(headers, QUESTION_VARIATIONS)
            if q_col_name:
                return r, headers[q_col_name], q_col_name
            else:
                return r, 2, None
    
    # Strategy 3: Fallback - look for Question column
    for r in range(1, scan_to + 1):
        headers = _header_index(ws, header_row=r)
        q = _find_column_case_insensitive(headers, QUESTION_VARIATIONS)
        if q:
            return r, headers[q], q
    
    # Last resort: assume header is in row 3 and questions in column 2
    return 3, 2, None


def _scan_for_question_format(ws) -> str:
    """Detect if a sheet is in row-based or column-based questionnaire format."""
    # Check if there's a "Question" or similar header
    for r in range(1, min(10, ws.max_row + 1)):
        headers = _header_index(ws, header_row=r)
        if _find_column_case_insensitive(headers, QUESTION_VARIATIONS):
            return "row-based"
    
    # Check for column-based format indicators
    column_indicators = [
        "vm hostname", "domain", "ip address", "application name",
        "server function", "operating system", "vcpu", "ram",
        "disks and size", "lun id", "server environment"
    ]
    
    for r in range(1, min(6, ws.max_row + 1)):
        headers = _header_index(ws, header_row=r)
        if headers:
            header_values = [str(h).lower() for h in headers.keys() if h]
            matches = sum(1 for ind in column_indicators if any(ind in hv for hv in header_values))
            if matches >= 3:
                return "column-based"
    
    return "row-based"


def _find_column_header_row(ws) -> int:
    """Heuristically find the header row in a column-based sheet (defaults to 3)."""
    for r in range(1, min(10, ws.max_row + 1)):
        non_empty_count = 0
        potential_headers = []
        
        for c in range(1, min(20, ws.max_column + 1)):
            cell_value = ws.cell(r, c).value
            if cell_value and str(cell_value).strip():
                non_empty_count += 1
                potential_headers.append(str(cell_value).lower())
        
        if non_empty_count >= 3:
            common_headers = [
                "hostname", "domain", "ip", "application", "server",
                "operating", "vcpu", "ram", "disk", "environment"
            ]
            matches = sum(1 for header in common_headers if any(header in ph for ph in potential_headers))
            if matches >= 2:
                return r
    
    return 3


def _find_first_blank_row(ws, header_row: int, max_scan: int = 20) -> int:
    """Find the first mostly-empty row after header_row (used for answers/confidence/provenance rows)."""
    start_row = header_row + 1
    max_row = min(header_row + max_scan, ws.max_row or header_row + 1)
    
    for r in range(start_row, max_row + 1):
        non_empty_count = 0
        total_cells = min(20, ws.max_column)
        
        for c in range(1, total_cells + 1):
            cell_value = ws.cell(r, c).value
            if cell_value and str(cell_value).strip():
                non_empty_count += 1
        
        if non_empty_count < (total_cells * 0.2):
            return r
    
    return header_row + 1

class QuestionnaireProcessorPlugin:
    """
    Excel processing plugin.
    - Loads the workbook template, detects per-sheet format (row-based vs column-based),
      and enumerates questions for the orchestrator.
    - Persists approved answers back into the workbook.
    """
    
    def __init__(self):
        self._loaded_questions: List[Dict] = []
        self._excel_path: str = ""
        self._column_maps: Dict[str, Dict[str, int | str]] = {}
        self._selected_sheets: List[str] = []
        self._sheet_formats: Dict[str, str] = {}  # Track format per sheet
        
    @kernel_function(description="Initialize questionnaire processing by loading Excel and preparing for Q&A")
    def initialize_questionnaire(
        self,
        path: Annotated[str, "Local filesystem path to the Excel workbook"],
        sheets_csv: Annotated[str, "Comma-separated sheet names to process; allow multiple; if empty, uses EXCEL_SHEETS env or active sheet"] = "",
    ) -> Annotated[str, "Status message with question count"]:
        """
        ENTRYPOINT: Call this first to prepare the workbook.
        - Creates a clean working copy of the Excel file
        - Detects sheet format and ensures required headers/rows exist
        - Builds an internal list of questions with positions (RowIndex/ColumnIndex)
        
        Args:
            path: Local path to the questionnaire workbook
            sheets_csv: Optional comma-separated sheet list; defaults to EXCEL_SHEETS or active sheet
        Returns:
            A status message with the number of questions and sheets processed
        """
        try:
            clean_path = materialize_clean_xlsx(path)
            wb = openpyxl.load_workbook(clean_path)
            # Prefer explicit argument, else fallback to environment variable EXCEL_SHEETS, else active sheet
            env_csv = (os.getenv("EXCEL_SHEETS") or "").strip()
            sheets = _parse_sheets_csv(sheets_csv or env_csv) or [wb.active.title]
            self._selected_sheets = sheets

            self._loaded_questions = []
            self._column_maps = {}
            self._sheet_formats = {}

            changed = False
            for sheet_name in sheets:
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                
                # Determine format
                format_type = _scan_for_question_format(ws)
                self._sheet_formats[sheet_name] = format_type
                
                if format_type == "row-based":
                    # Row-based: questions in rows; ensure Answer/Confidence/Provenance columns
                    header_row, q_col_idx, _ = _scan_for_question_header(ws)
                    headers = _header_index(ws, header_row=header_row)

                    # Ensure Answer column exists
                    original_headers_count = len(headers)
                    headers, answer_col_name = _find_or_create_column(ws, headers, ANSWER_VARIATIONS, "Response", header_row)
                    if len(headers) > original_headers_count:
                        changed = True
                    answer_col_idx = headers[answer_col_name]

                    # Optional guidance column detection
                    guidance_col_name = _find_column_case_insensitive(headers, GUIDANCE_VARIATIONS)
                    guidance_col_idx = headers.get(guidance_col_name) if guidance_col_name else None

                    # Ensure Confidence column exists
                    original_headers_count = len(headers)
                    headers, confidence_col_name = _find_or_create_column(ws, headers, ["Confidence"], "Confidence", header_row)
                    if len(headers) > original_headers_count:
                        changed = True
                    confidence_col_idx = headers[confidence_col_name]
                    
                    # Ensure Provenance column exists
                    original_headers_count = len(headers)
                    headers, provenance_col_name = _find_or_create_column(ws, headers, ["Provenance"], "Provenance", header_row)
                    if len(headers) > original_headers_count:
                        changed = True
                    provenance_col_idx = headers[provenance_col_name]

                    if not q_col_idx:
                        q_col_idx = 2

                    self._column_maps[sheet_name] = {
                        "format": "row-based",
                        "header_row": header_row,
                        "question_col": q_col_idx,
                        "answer_col": answer_col_idx,
                        "guidance_col": guidance_col_idx or 0,
                        "confidence_col": confidence_col_idx,
                        "provenance_col": provenance_col_idx,
                    }

                    # Load questions - start from the row after the header
                    start_row = header_row + 1
                    max_row = ws.max_row or start_row
                    
                    for r in range(start_row, max_row + 1):
                        qv = ws.cell(r, q_col_idx).value
                        if qv is None or str(qv).strip() == "":
                            continue
                        guidance_val = ws.cell(r, guidance_col_idx).value if guidance_col_idx else None
                        self._loaded_questions.append({
                            "RowIndex": r,
                            "ColumnIndex": 0,  # Not used for row-based
                            "SheetName": sheet_name,
                            "Question": str(qv),
                            "Guidance": str(guidance_val) if guidance_val is not None else "",
                            "Format": "row-based"
                        })
                
                else:  # column-based
                    # Column-based: questions in columns; answers/confidence/provenance are rows
                    # Find the actual header row
                    header_row = _find_column_header_row(ws)
                    
                    # Find the first blank row below the header for answers
                    answer_row = _find_first_blank_row(ws, header_row)
                    
                    # Find or create Confidence and Provenance rows (after answer row)
                    confidence_row = _find_first_blank_row(ws, answer_row)
                    if confidence_row == answer_row:
                        confidence_row = answer_row + 1
                    
                    provenance_row = _find_first_blank_row(ws, confidence_row)
                    if provenance_row == confidence_row:
                        provenance_row = confidence_row + 1
                    
                    # Add row headers for Confidence and Provenance
                    if ws.cell(confidence_row, 1).value != "Confidence":
                        ws.cell(confidence_row, 1, value="Confidence")
                        changed = True
                    
                    if ws.cell(provenance_row, 1).value != "Provenance":
                        ws.cell(provenance_row, 1, value="Provenance")
                        changed = True
                    
                    self._column_maps[sheet_name] = {
                        "format": "column-based",
                        "header_row": header_row,
                        "answer_row": answer_row,
                        "confidence_row": confidence_row,
                        "provenance_row": provenance_row,
                    }
                    
                    # Load ALL columns as questions (including metadata columns)
                    for col in range(1, ws.max_column + 1):
                        header_val = ws.cell(header_row, col).value
                        if header_val and str(header_val).strip():
                            header_str = str(header_val).strip()
                            if not header_str:
                                continue
                            self._loaded_questions.append({
                                "RowIndex": answer_row,  # Answer will go in the first blank row
                                "ColumnIndex": col,
                                "SheetName": sheet_name,
                                "Question": header_str,
                                "Guidance": "",  # No guidance in column-based format
                                "Format": "column-based"
                            })

            if changed:
                wb.save(clean_path)
            
            self._excel_path = clean_path
            
            return f"Loaded {len(self._loaded_questions)} questions from {len(self._column_maps)} sheet(s)."
        except Exception as e:
            import traceback
            traceback.print_exc()
            return f"Error: {e}"

    @kernel_function(description="Save approved answers back to Excel")
    def persist_answers(
        self,
        updates_json: Annotated[str, "JSON array of updates with RowIndex, ColumnIndex, SheetName, Answer, Confidence, Provenance"],
    ) -> Annotated[str, "Status message"]:
        """
        Persist a set of approved answers into the working Excel file.
        - Row-based: writes to Answer/Confidence/Provenance columns at the question row
        - Column-based: writes to Answer/Confidence/Provenance rows at the question column
        
        Args:
            updates_json: JSON array of updates produced by ReviewPlugin.get_updates_json
        Returns:
            A status message indicating how many updates were applied
        """
        if not self._excel_path:
            return "Error: Excel not initialized"
        try:
            updates = json.loads(updates_json) if isinstance(updates_json, str) else updates_json
            if not isinstance(updates, list):
                return "Error: updates_json must be a JSON array"

            wb = openpyxl.load_workbook(self._excel_path)

            # Helper to write to data cells (handles merged cells properly)
            def write_data_cell(ws, row, col, value):
                """Write to a specific cell, unmerging if the target is inside a merged region."""
                if col <= 0 or row <= 0:
                    return  # Invalid column/row
                merged_ranges_to_unmerge = []
                for merged_range in ws.merged_cells.ranges:
                    if row >= merged_range.min_row and row <= merged_range.max_row and \
                       col >= merged_range.min_col and col <= merged_range.max_col:
                        merged_ranges_to_unmerge.append(str(merged_range))
                for range_str in merged_ranges_to_unmerge:
                    ws.unmerge_cells(range_str)
                ws.cell(row, col, value=value)

            updates_applied = 0
            for upd in updates:
                sheet = upd.get("SheetName")
                if sheet not in wb.sheetnames or sheet not in self._column_maps:
                    continue
                ws = wb[sheet]
                col_info = self._column_maps[sheet]
                
                if col_info.get("format") == "row-based":
                    # Existing row-based logic
                    row = int(upd.get("RowIndex")) if upd.get("RowIndex") is not None else None
                    if row is None:
                        continue
                    answer = upd.get("Answer", "")
                    confidence = upd.get("Confidence", "")
                    provenance = upd.get("Provenance", "")
                    header_row = col_info.get("header_row", 1)
                    # Confidence header and value
                    confidence_col = col_info.get("confidence_col", 0)
                    if confidence_col > 0:
                        if ws.cell(header_row, confidence_col).value is None:
                            ws.cell(header_row, confidence_col, value="Confidence")
                        write_data_cell(ws, row, confidence_col, confidence)
                    # Provenance header and value
                    provenance_col = col_info.get("provenance_col", 0)
                    if provenance_col > 0:
                        if ws.cell(header_row, provenance_col).value is None:
                            ws.cell(header_row, provenance_col, value="Provenance")
                        write_data_cell(ws, row, provenance_col, provenance)
                    # Answer
                    answer_col = col_info.get("answer_col", 0)
                    if answer_col > 0:
                        write_data_cell(ws, row, answer_col, answer)
                    updates_applied += 1
                
                else:  # column-based
                    # Resolve column index via ColumnIndex or fallback to Question match
                    col_idx = int(upd.get("ColumnIndex", 0)) if upd.get("ColumnIndex") else 0
                    if col_idx <= 0:
                        question_text = upd.get("Question", "")
                        if question_text:
                            for loaded_q in self._loaded_questions:
                                if (loaded_q.get("SheetName") == sheet and 
                                    loaded_q.get("Question") == question_text and
                                    loaded_q.get("Format") == "column-based"):
                                    col_idx = loaded_q.get("ColumnIndex", 0)
                                    break
                    if col_idx <= 0:
                        continue
                    answer = upd.get("Answer", "")
                    confidence = upd.get("Confidence", "")
                    provenance = upd.get("Provenance", "")
                    answer_row = col_info.get("answer_row", 4)
                    confidence_row = col_info.get("confidence_row", answer_row + 1)
                    provenance_row = col_info.get("provenance_row", answer_row + 2)
                    write_data_cell(ws, answer_row, col_idx, answer)
                    if confidence_row > 0:
                        write_data_cell(ws, confidence_row, col_idx, confidence)
                    if provenance_row > 0:
                        write_data_cell(ws, provenance_row, col_idx, provenance)
                    updates_applied += 1

            wb.save(self._excel_path)
            return f"Saved {updates_applied} answers to Excel."
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return f"Error: {e}"

    @kernel_function(description="Get current Excel file path for upload")
    def get_excel_path(self) -> Annotated[str, "Path to the processed Excel file"]:
        """Return the path of the working Excel file created during initialize_questionnaire."""
        return self._excel_path or ""

    @kernel_function(description="Return loaded questions as JSON array")
    def get_loaded_questions(self) -> Annotated[str, "JSON array of loaded questions with RowIndex, SheetName, Question, Guidance"]:
        """Return a JSON array of questions and their positions gathered during initialization."""
        return json.dumps(self._loaded_questions or [])

    @kernel_function(description="Get per-sheet question map with ordinal numbers")
    def get_question_map(self) -> Annotated[str, "JSON { sheetName: [ {Number, RowIndex, Question} ] }"]:
        """Build an ordinal list per sheet from the loaded questions (useful for interactive review)."""
        per_sheet: Dict[str, List[Dict]] = {}
        for q in self._loaded_questions or []:
            s = q.get("SheetName", "")
            if not s:
                continue
            per_sheet.setdefault(s, [])
            per_sheet[s].append({
                "RowIndex": int(q.get("RowIndex", 0)),
                "Question": q.get("Question", "")
            })
        # annotate ordinals
        for s, items in per_sheet.items():
            for i, item in enumerate(items, start=1):
                item["Number"] = i
        return json.dumps(per_sheet)

    @kernel_function(description="Resolve a question number to its row index within the given sheet")
    def resolve_question_number(
        self,
        sheet: Annotated[str, "Sheet name"],
        number: Annotated[int, "1-based question number within the sheet"],
    ) -> Annotated[str, "JSON with { RowIndex, Question } or {} if not found"]:
        """Resolve a per-sheet ordinal to its RowIndex and Question text."""
        if not sheet or number is None:
            return "{}"
        # Build on demand from current loaded questions
        filtered = [q for q in (self._loaded_questions or []) if q.get("SheetName") == sheet]
        if number < 1 or number > len(filtered):
            return "{}"
        q = filtered[number - 1]
        return json.dumps({
            "RowIndex": int(q.get("RowIndex", 0)),
            "Question": q.get("Question", "")
        })


class BlobPlugin:
    """Lightweight plugin for Azure Blob operations used by the agent."""
    
    @kernel_function(description="Download a blob from Azure Storage")
    def download_blob(
        self,
        container: Annotated[str, "Azure Storage container name"],
        blob_name: Annotated[str, "Name of the blob to download"],
    ) -> Annotated[str, "Local file path to the downloaded blob"]:
        """Download a blob and return the local file path."""
        return blob_utils.download_blob(container, blob_name)
    
    @kernel_function(description="Upload a file to Azure Storage")
    def upload_file(
        self,
        container: Annotated[str, "Azure Storage container name"],
        local_path: Annotated[str, "Local filesystem path to upload"],
        dest_blob_name: Annotated[str, "Destination blob name"] = "",
    ) -> Annotated[str, "URL of the uploaded blob"]:
        """Upload a local file to blob storage and return the URL."""
        return blob_utils.upload_file(container, local_path, dest_blob_name or None)
    
    @kernel_function(description="List blobs in a container")
    def list_blobs(
        self,
        container: Annotated[str, "Azure Storage container name"],
        prefix: Annotated[str, "Optional prefix filter"] = "",
    ) -> Annotated[str, "JSON array of blob names"]:
        """List blob names in the specified container (optionally filtered by prefix)."""
        blobs = blob_utils.list_blobs(container, prefix or None)
        return json.dumps(blobs)


class ReviewPlugin:
    """Transient store for answer proposals and review utilities."""
    
    def __init__(self):
        self._proposals: List[Dict] = []
        self._title: str = ""
    
    @kernel_function(description="Get current proposal count for debugging")
    def get_proposal_count(self) -> Annotated[str, "Number of proposals currently stored"]:
        """Return the number of proposals currently cached in memory."""
        return str(len(self._proposals))
    
    @kernel_function(description="Add proposals from processing (upsert by SheetName+RowIndex)")
    def add_proposals(
        self,
        proposals_json: Annotated[str, "JSON with proposals array or object with 'proposals' key"],
    ) -> Annotated[str, "Status with total count"]:
        """Upsert proposals keyed by (SheetName, RowIndex, ColumnIndex) and keep the latest values."""
        try:
            data = json.loads(proposals_json) if isinstance(proposals_json, str) else proposals_json
            items = data.get("proposals") if isinstance(data, dict) else data
            if not isinstance(items, list):
                return "Error: proposals_json must be an array or { 'proposals': [] }"
            
            # upsert by (SheetName, RowIndex, ColumnIndex)
            by_key: Dict[Tuple[str, int, int], Dict] = {}
            for p in self._proposals:
                key = (p.get("SheetName", ""), int(p.get("RowIndex", 0)), int(p.get("ColumnIndex", 0)))
                by_key[key] = p
            
            for it in items:
                sheet = it.get("SheetName", "")
                row = int(it.get("RowIndex", 0))
                col = int(it.get("ColumnIndex", 0))
                key = (sheet, row, col)
                by_key[key] = {
                    "SheetName": sheet,
                    "RowIndex": row,
                    "ColumnIndex": col,
                    "Question": it.get("Question", ""),
                    "Answer": it.get("Answer", ""),
                    "Confidence": it.get("Confidence", "Unknown"),
                    "Provenance": it.get("Provenance", ""),
                }
            self._proposals = list(by_key.values())
            return f"Total proposals: {len(self._proposals)}"
        except Exception as e:
            return f"Error: {e}"
    
    @kernel_function(description="Render proposals as Markdown table")
    def render_markdown(self) -> Annotated[str, "Markdown table"]:
        """Render the current proposals as a Markdown table (for optional preview)."""
        if not self._proposals:
            return "No proposals to review."
        # sort for stable output
        rows = sorted(self._proposals, key=lambda x: (x.get("SheetName", ""), int(x.get("RowIndex", 0))))
        header = "| # | Sheet | Row | Question | Answer | Confidence | Provenance |\n|---|---|---:|---|---|---|---|"
        lines = [header]
        for i, p in enumerate(rows, start=1):
            q = (p.get("Question", "") or "").replace("\n", " ")
            a = (p.get("Answer", "") or "").replace("\n", " ")
            conf = p.get("Confidence", "Unknown") or "Unknown"
            prov = (p.get("Provenance", "") or "").replace("\n", " ")
            lines.append(f"| {i} | {p.get('SheetName','')} | {p.get('RowIndex','')} | {q} | {a} | {conf} | {prov} |")
        return "\n".join(lines)
    
    @kernel_function(description="Get approved proposals as JSON (RowIndex, SheetName, Answer, Confidence, Provenance)")
    def get_updates_json(self) -> Annotated[str, "JSON array of updates"]:
        """Return all proposals as a JSON array for persistence into Excel."""
        updates = []
        for p in self._proposals:
            updates.append({
                "SheetName": p.get("SheetName", ""),
                "RowIndex": p.get("RowIndex", 0),
                "ColumnIndex": p.get("ColumnIndex", 0),
                "Question": p.get("Question", ""),
                "Answer": p.get("Answer", ""),
                "Confidence": p.get("Confidence", "Unknown"),
                "Provenance": p.get("Provenance", ""),
            })
        return json.dumps(updates)
    
    @kernel_function(description="Clear all proposals")
    def clear(self) -> Annotated[str, "Status"]:
        """Clear all proposals from memory."""
        self._proposals = []
        self._title = ""
        return "Cleared"

    @kernel_function(description="Update/set an answer for a given Sheet+Row; sets Provenance to 'User_filled' by default")
    def set_answer(
        self,
        sheet: Annotated[str, "Sheet name"],
        row_index: Annotated[int, "Row index of the question row"],
        answer: Annotated[str, "New answer text"],
        confidence: Annotated[str, "Optional confidence label"] = "High",
        provenance: Annotated[str, "Optional provenance"] = "User_filled",
    ) -> Annotated[str, "Status"]:
        """Upsert a manual answer for review (useful during interactive completion)."""
        try:
            if not sheet or not row_index:
                return "Error: sheet and row_index are required"
            # upsert on (SheetName, RowIndex)
            key = (sheet, int(row_index))
            by_key: Dict[Tuple[str, int], Dict] = {}
            for p in self._proposals:
                by_key[(p.get("SheetName",""), int(p.get("RowIndex",0)))] = p
            # keep existing question text if present
            existing = by_key.get(key, {"SheetName": sheet, "RowIndex": int(row_index)})
            existing["Answer"] = answer or ""
            existing["Confidence"] = confidence or "High"
            existing["Provenance"] = provenance or "User_filled"
            if "Question" not in existing or not existing["Question"]:
                existing["Question"] = ""
            by_key[key] = existing
            self._proposals = list(by_key.values())
            return f"Updated {sheet}#{row_index}"
        except Exception as e:
            return f"Error: {e}"

    @kernel_function(description="List unanswered proposals (Answer empty) for review")
    def get_unanswered(self) -> Annotated[str, "JSON array of {SheetName, RowIndex, Question}"]:
        """Return proposals that still have an empty Answer field."""
        items = []
        for p in self._proposals:
            if not (p.get("Answer") or "").strip():
                items.append({
                    "SheetName": p.get("SheetName",""),
                    "RowIndex": int(p.get("RowIndex",0)),
                    "Question": p.get("Question",""),
                })
        return json.dumps(items)

    @kernel_function(description="Return count of unanswered proposals")
    def get_unanswered_count(self) -> Annotated[str, "Number of unanswered entries"]:
        """Return the count of proposals whose Answer is empty."""
        cnt = 0
        for p in self._proposals:
            if not (p.get("Answer") or "").strip():
                cnt += 1
        return str(cnt)

    @kernel_function(description="Fix missing column headers in Excel")
    def fix_headers(self) -> Annotated[str, "Status message"]:
        """Ensure Confidence/Provenance headers exist in row-based sheets and save the workbook if fixed."""
        if not self._excel_path or not self._column_maps:
            return "Error: Excel not initialized"
        
        try:
            wb = openpyxl.load_workbook(self._excel_path)
            fixed_count = 0
            
            for sheet_name, col_info in self._column_maps.items():
                if sheet_name not in wb.sheetnames:
                    continue
                    
                ws = wb[sheet_name]
                
                if col_info.get("format") == "row-based":
                    header_row = col_info.get("header_row", 1)
                    
                    # Fix Confidence header
                    confidence_col = col_info.get("confidence_col", 0)
                    if confidence_col > 0:
                        if ws.cell(header_row, confidence_col).value is None:
                            ws.cell(header_row, confidence_col, value="Confidence")
                            fixed_count += 1
                    
                    # Fix Provenance header
                    provenance_col = col_info.get("provenance_col", 0)
                    if provenance_col > 0:
                        if ws.cell(header_row, provenance_col).value is None:
                            ws.cell(header_row, provenance_col, value="Provenance")
                            fixed_count += 1
            
            if fixed_count > 0:
                wb.save(self._excel_path)
            
            return f"Fixed {fixed_count} missing headers"
            
        except Exception as e:
            return f"Error fixing headers: {e}"