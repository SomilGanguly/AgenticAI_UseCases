import openpyxl
from typing import List, Dict, Optional, Tuple, Set
import re

# Column name variations to search for (case-insensitive)
QUESTION_VARIATIONS = ["Question", "Questions", "question", "questions", "Query", "query", "Q", "q"]
GUIDANCE_VARIATIONS = ["Guidance", "guidance", "Examples", "examples", "Instructions", "instructions", "Help", "help", "Guide", "guide"]
ANSWER_VARIATIONS = ["Response", "Responses", "response", "responses", "Answer", "Answers", "answer", "answers", "Reply", "reply", "A", "a"]

# Additional tracking columns
TRACKING_COLUMNS = ["Confidence", "Provenance"]

def _parse_sheet_names(sheet_names_str: Optional[str]) -> Optional[List[str]]:
    """Parse comma-separated sheet names from environment variable."""
    if not sheet_names_str:
        return None
    # Split by comma and strip whitespace
    return [name.strip() for name in sheet_names_str.split(',') if name.strip()]

def _get_sheets_to_process(wb: openpyxl.Workbook, target_sheets: Optional[List[str]] = None) -> List[str]:
    """Get list of sheet names to process."""
    all_sheets = wb.sheetnames
    
    if not target_sheets:
        # If no specific sheets requested, use the active sheet only
        return [wb.active.title]
    
    # Validate requested sheets exist
    valid_sheets = []
    invalid_sheets = []
    
    for sheet_name in target_sheets:
        if sheet_name in all_sheets:
            valid_sheets.append(sheet_name)
        else:
            invalid_sheets.append(sheet_name)
    
    if invalid_sheets:
        print(f"⚠️  Warning: The following sheets were not found in the workbook: {', '.join(invalid_sheets)}")
        print(f"   Available sheets: {', '.join(all_sheets)}")
    
    if not valid_sheets:
        raise ValueError(f"None of the requested sheets found in workbook. Requested: {', '.join(target_sheets)}, Available: {', '.join(all_sheets)}")
    
    return valid_sheets

def _find_column_case_insensitive(headers: Dict[str, int], variations: List[str]) -> Optional[str]:
    """Find the first matching column name from variations (case-insensitive)."""
    # First try exact match
    for var in variations:
        if var in headers:
            return var
    
    # Then try case-insensitive match
    headers_lower = {k.lower(): k for k in headers.keys() if k is not None}
    for var in variations:
        if var.lower() in headers_lower:
            return headers_lower[var.lower()]
    
    # Try partial match (contains)
    for header in headers.keys():
        if header is None:
            continue
        header_lower = header.lower()
        for var in variations:
            if var.lower() in header_lower or header_lower in var.lower():
                return header
    
    return None

def _find_or_create_column(ws, headers: Dict[str, int], variations: List[str], default_name: str, header_row: int = 1) -> Tuple[Dict[str, int], str]:
    """Find a column from variations or create it with default name at header_row."""
    found = _find_column_case_insensitive(headers, variations)
    if found:
        return headers, found
    
    # Create new column at the end of the header row
    new_col_idx = len(headers) + 1
    ws.cell(header_row, new_col_idx, value=default_name)
    headers[default_name] = new_col_idx
    return headers, default_name

def _header_index(ws, header_row: int = 1) -> Dict[str, int]:
    """Create a mapping of column names to their 1-based indices for a given header row."""
    headers = {}
    # Protect against asking for a row beyond sheet
    if header_row < 1 or header_row > (ws.max_row or 1):
        return headers
    # ws[header_row] returns the row tuple
    for idx, cell in enumerate(ws[header_row], start=1):
        if cell.value is not None:
            header_name = str(cell.value).strip()
            if header_name:
                headers[header_name] = idx
    return headers

def _find_question_column(ws, max_scan_rows: int = 7) -> Tuple[Optional[int], Optional[int], Optional[str]]:
    """
    Scan the first `max_scan_rows` rows for a cell whose text matches a Question variation.
    Returns (header_row, column_index, header_text) or (None, None, None) if not found.
    """
    q_variants = {v.lower() for v in QUESTION_VARIATIONS}
    max_row = max(1, ws.max_row or 1)
    scan_to = min(max_scan_rows, max_row)
    for r in range(1, scan_to + 1):
        # iterate cells in row r
        for col_index, cell in enumerate(ws[r], start=1):
            v = cell.value
            if v is None:
                continue
            txt = str(v).strip()
            if not txt:
                continue
            if txt.lower() in q_variants:
                return r, col_index, txt
    return None, None, None

def ensure_columns(path: str, target_sheets: Optional[List[str]] = None) -> Dict[str, Dict[str, str]]:
    """Ensure required columns exist in specified sheets and return column name mappings per sheet.
    This respects a detected header row (scans top rows for 'Question' if not on row 1)."""
    wb = openpyxl.load_workbook(path)
    sheets_to_process = _get_sheets_to_process(wb, target_sheets)
    
    sheet_column_maps = {}
    changed = False
    MAX_SCAN_ROWS = 7
    
    for sheet_name in sheets_to_process:
        ws = wb[sheet_name]
        # Try header row 1 first
        header_row_used = 1
        headers = _header_index(ws, header_row=1)
        question_col = _find_column_case_insensitive(headers, QUESTION_VARIATIONS) if headers else None

        # If not found on row 1, scan first MAX_SCAN_ROWS to find the Question header
        if not question_col:
            hr, q_col_idx, header_text = _find_question_column(ws, max_scan_rows=MAX_SCAN_ROWS)
            if hr and q_col_idx:
                header_row_used = hr
                headers = _header_index(ws, header_row=header_row_used)
                question_col = header_text
            else:
                print(f"\n📋 Processing sheet '{sheet_name}'")
                print(f"   Found columns: {list(headers.keys()) if headers else []}")
                print(f"   ⚠️  Skipping sheet '{sheet_name}': No question column found in top {MAX_SCAN_ROWS} rows")
                continue

        print(f"\n📋 Processing sheet '{sheet_name}' (header row: {header_row_used})")
        print(f"   Found columns: {list(headers.keys())}")

        # Find or create answer column on the detected header row
        original_headers = headers.copy()
        headers, answer_col = _find_or_create_column(ws, headers, ANSWER_VARIATIONS, "Response", header_row=header_row_used)
        if headers != original_headers:
            changed = True

        # Find guidance column (optional)
        guidance_col = _find_column_case_insensitive(headers, GUIDANCE_VARIATIONS)

        # Ensure tracking columns exist at header_row_used
        for col in TRACKING_COLUMNS:
            if col not in headers:
                ws.cell(header_row_used, len(headers) + 1, value=col)
                headers[col] = len(headers) + 1
                changed = True

        # Store column mappings for this sheet, include header_row_used
        sheet_column_maps[sheet_name] = {
            "question": question_col,
            "guidance": guidance_col,
            "answer": answer_col,
            "confidence": "Confidence",
            "provenance": "Provenance",
            "header_row": header_row_used
        }
    
    if changed:
        wb.save(path)
    
    return sheet_column_maps

def load_questions(path: str, target_sheets: Optional[List[str]] = None) -> Tuple[List[Dict], Dict[str, Dict[str, str]]]:
    """Load questions from specified sheets, returning RowIndex, SheetName, and column mappings."""
    wb = openpyxl.load_workbook(path)
    sheets_to_process = _get_sheets_to_process(wb, target_sheets)
    
    all_rows: List[Dict] = []
    sheet_column_maps = {}
    MAX_SCAN_ROWS = 7
    
    for sheet_name in sheets_to_process:
        ws = wb[sheet_name]

        # Try header row 1 first
        headers = _header_index(ws, header_row=1)
        question_col_name = _find_column_case_insensitive(headers, QUESTION_VARIATIONS) if headers else None
        header_row_used = 1

        # If no question header on row1, scan rows 1..MAX_SCAN_ROWS for a Question header
        if not question_col_name:
            hr, q_col_idx, header_text = _find_question_column(ws, max_scan_rows=MAX_SCAN_ROWS)
            if hr and q_col_idx:
                header_row_used = hr
                # Build headers mapping for that header row
                headers = _header_index(ws, header_row=header_row_used)
                question_col_name = header_text
            else:
                print(f"\n📋 Loading from sheet '{sheet_name}'")
                print(f"   Columns: {list(headers.keys()) if headers else []}")
                print(f"   ⚠️  Skipping sheet '{sheet_name}': No question column found in top {MAX_SCAN_ROWS} rows")
                continue

        print(f"\n📋 Loading from sheet '{sheet_name}' (header row: {header_row_used})")
        print(f"   Columns: {list(headers.keys())}")

        # Find other columns relative to the detected header row
        guidance_col_name = _find_column_case_insensitive(headers, GUIDANCE_VARIATIONS)
        answer_col_name = _find_column_case_insensitive(headers, ANSWER_VARIATIONS)
        
        print(f"   Using columns - Question: '{question_col_name}', Guidance: '{guidance_col_name}', Answer: '{answer_col_name}'")
        
        # Store column mappings for this sheet, include header_row
        sheet_column_maps[sheet_name] = {
            "question": question_col_name,
            "guidance": guidance_col_name,
            "answer": answer_col_name,
            "confidence": "Confidence",
            "provenance": "Provenance",
            "header_row": header_row_used
        }
        
        rows: List[Dict] = []
        q_col = headers.get(question_col_name) if question_col_name else None
        g_col = headers.get(guidance_col_name) if guidance_col_name else None
        cat_col = headers.get("Category")
        pri_col = headers.get("Priority")
        
        # Iterate data rows starting after header_row_used
        for offset_idx, row in enumerate(ws.iter_rows(min_row=header_row_used + 1, values_only=True), start=header_row_used + 1):
            # row is a tuple of values aligned to columns
            if not q_col:
                continue
            question = row[q_col - 1] if row and len(row) >= q_col else None
            if question is None or str(question).strip() == "":
                continue
                
            item = {
                "RowIndex": offset_idx,
                "SheetName": sheet_name,  # Add sheet name to track source
                "Question": str(question).strip(),
            }
            
            # Add guidance if available
            if g_col and len(row) >= g_col:
                guidance = row[g_col - 1]
                if guidance and str(guidance).strip():
                    item["Guidance"] = str(guidance).strip()
            
            # Add optional columns
            if cat_col and len(row) >= cat_col:
                item["Category"] = row[cat_col - 1]
            if pri_col and len(row) >= pri_col:
                item["Priority"] = row[pri_col - 1]
                
            rows.append(item)
        
        print(f"   Loaded {len(rows)} questions from sheet '{sheet_name}'")
        all_rows.extend(rows)
    
    print(f"\n✅ Total questions loaded: {len(all_rows)} from {len(sheet_column_maps)} sheet(s)")
    
    return all_rows, sheet_column_maps

def _get_top_left_of_merged(ws, row: int, col: int) -> (int, int):
    """If (row,col) is inside a merged cell, return the top-left (min_row, min_col) of that merged range.
    Otherwise return (row, col)."""
    for cr in ws.merged_cells.ranges:
        if cr.min_row <= row <= cr.max_row and cr.min_col <= col <= cr.max_col:
            return cr.min_row, cr.min_col
    return row, col

def update_answers(path: str, updates: List[Dict], sheet_column_maps: Optional[Dict[str, Dict[str, str]]] = None, target_sheets: Optional[List[str]] = None) -> None:
    """Update rows by RowIndex and SheetName, writing answers and tracking information."""
    wb = openpyxl.load_workbook(path)
    
    # If column_map not provided, detect columns (ensure_columns will detect header rows)
    if not sheet_column_maps:
        sheet_column_maps = ensure_columns(path, target_sheets)
    
    # Group updates by sheet
    updates_by_sheet: Dict[str, List[Dict]] = {}
    for update in updates:
        sheet_name = update.get("SheetName")
        if sheet_name:
            if sheet_name not in updates_by_sheet:
                updates_by_sheet[sheet_name] = []
            updates_by_sheet[sheet_name].append(update)
    
    # Process each sheet
    for sheet_name, sheet_updates in updates_by_sheet.items():
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  Warning: Sheet '{sheet_name}' not found in workbook")
            continue
            
        ws = wb[sheet_name]
        column_map = sheet_column_maps.get(sheet_name, {})
        header_row = int(column_map.get("header_row", 1))
        headers = _header_index(ws, header_row=header_row)
        
        # Ensure all required columns exist on header_row
        for col_type, col_name in column_map.items():
            # skip metadata keys that aren't column names
            if col_type == "header_row":
                continue
            if col_name and col_name not in headers:
                new_col_idx = len(headers) + 1
                ws.cell(header_row, new_col_idx, value=col_name)
                headers = _header_index(ws, header_row=header_row)  # refresh after adding
        
        for u in sheet_updates:
            row_idx = u.get("RowIndex")
            if not row_idx:
                continue
                
            # Write answer to the appropriate column (handle merged cells)
            ans_col_key = column_map.get("answer")
            if "Answer" in u and ans_col_key and ans_col_key in headers:
                col_idx = headers[ans_col_key]
                top_row, top_col = _get_top_left_of_merged(ws, row_idx, col_idx)
                ws.cell(top_row, top_col).value = u.get("Answer")
                
            # Write tracking information (handle merged cells similarly)
            conf_col_key = column_map.get("confidence")
            prov_col_key = column_map.get("provenance")
            if "Confidence" in u and conf_col_key and conf_col_key in headers:
                col_idx = headers[conf_col_key]
                top_row, top_col = _get_top_left_of_merged(ws, row_idx, col_idx)
                ws.cell(top_row, top_col).value = u.get("Confidence")
            if "Provenance" in u and prov_col_key and prov_col_key in headers:
                col_idx = headers[prov_col_key]
                top_row, top_col = _get_top_left_of_merged(ws, row_idx, col_idx)
                ws.cell(top_row, top_col).value = u.get("Provenance")
    
    wb.save(path)